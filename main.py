from flask import Flask, request
import requests
import openpyxl
from openpyxl import Workbook, load_workbook
from urllib.parse import urlparse


htmlcontent=""
def download_file_from_google_drive(id, destination, htmlcontent):
    URL = "https://docs.google.com/uc?export=download"

    session = requests.Session()

    response = session.get(URL, params = { 'id' : id }, stream = True)
    token = get_confirm_token(response) 
    

    if token:
        params = { 'id' : id, 'confirm' : token }
        response = session.get(URL, params = params, stream = True)

    res=save_response_content(response, destination)   
    return res

def get_confirm_token(response):
    for key, value in response.cookies.items():
        if key.startswith('download_warning'):
            return value

    return None

def save_response_content(response, destination):
    global htmlcontent
    CHUNK_SIZE = 32768

    with open(destination, "wb") as f:
        for chunk in response.iter_content(CHUNK_SIZE):
            if chunk: # filter out keep-alive new chunks
                htmlcontent+=chunk.decode("utf-8") 

    return htmlcontent





app = Flask(__name__)

@app.route("/" , )
def hello_world():
    global htmlcontent
    # data=request.get_json()
    fileId=request.args.get("fileId")
    print(fileId)
        
    o = urlparse(request.base_url)
    print( "this is host id", o.hostname)

    
    wb=load_workbook("driveUrl.xlsx")

    ws=wb.active
    
    for i in range(1, ws.max_row):
        file_id=ws.cell(row=i,column=1).value
        fileMaskUrl=ws.cell(row=i, column=2).value
        if fileId==fileMaskUrl:
            print(file_id)
            break


    htmlcontent=""
    destination = './new.html'
    responseData=download_file_from_google_drive(file_id, destination, htmlcontent)
    return responseData
if __name__ == "__main__":
    app.run(debug=True, port=5000)
